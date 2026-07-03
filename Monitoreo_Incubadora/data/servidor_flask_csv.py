# servidor_flask_csv.py (CSV version con claves ordenadas)
from flask import Flask, jsonify, request
from flask_cors import CORS
from statistics import mean, median, stdev
import os, csv, glob, re
from flask import send_file
import io
from datetime import datetime
from auditoria import (
    actualizar_actividad,
    cargar_historial,
    cerrar_sesion,
    crear_sesion,
    sesion_activa,
    validar_credenciales,
)

app = Flask(__name__)
CORS(app)

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "scripts", "data"))
MAX_ROWS = 2160  # hasta 6 h con muestras cada 10 s


@app.route("/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    usuario = validar_credenciales(data.get("email"), data.get("password"))
    if not usuario:
        return jsonify({"ok": False, "error": "Credenciales inválidas"}), 401

    sesion = crear_sesion(
        usuario["email"],
        nombre=usuario["nombre"],
        ip=request.remote_addr,
        user_agent=request.headers.get("User-Agent"),
    )
    return jsonify({
        "ok": True,
        "session_id": sesion["session_id"],
        "email": sesion["email"],
        "nombre": sesion["nombre"],
        "inicio": sesion["inicio"],
    }), 200


@app.route("/session", methods=["GET", "POST"])
def session_status():
    data = request.get_json(silent=True) or {}
    session_id = (
        request.headers.get("X-Session-Id")
        or request.args.get("session_id")
        or data.get("session_id")
    )
    sesion = actualizar_actividad(session_id)
    if not sesion:
        return jsonify({"ok": False, "error": "Sesión inválida"}), 401
    return jsonify({
        "ok": True,
        "email": sesion["email"],
        "nombre": sesion.get("nombre", sesion["email"]),
        "inicio": sesion["inicio"],
        "ultima_actividad": sesion["ultima_actividad"],
    }), 200


@app.route("/logout", methods=["POST"])
def logout():
    data = request.get_json(silent=True) or {}
    session_id = request.headers.get("X-Session-Id") or data.get("session_id")
    sesion = cerrar_sesion(session_id)
    if not sesion:
        return jsonify({"ok": False, "error": "Sesión inválida o ya cerrada"}), 400
    return jsonify({"ok": True, "fin": sesion["fin"]}), 200


@app.route("/historial_sesiones", methods=["GET"])
def historial_sesiones():
    return jsonify(cargar_historial()), 200

def calcular_estadisticas(valores):
    nums = []
    for v in valores:
        try:
            nums.append(float(v))
        except:
            pass
    if len(nums) < 2:
        return {"media": None, "mediana": None, "min": None, "max": None, "desvio": None}
    return {
        "media":   round(mean(nums), 2),
        "mediana": round(median(nums), 2),
        "min":     round(min(nums), 2),
        "max":     round(max(nums), 2),
        "desvio":  round(stdev(nums), 2)
    }

def read_csv_rows(path, max_rows=MAX_ROWS):
    rows = []
    if not os.path.exists(path):
        return rows
    with open(path, newline='', encoding='utf-8') as f:
        reader = list(csv.DictReader(f))
        rows = reader[:max_rows]  # primeras max_rows (asumo más recientes arriba)
    return rows

def construir_datos_parametro(filas, valor_key, ley_key, grafico_key):
    datos = []
    historial = []
    punto_actual = None

    for fila in filas:
        fecha = fila.get("fecha")
        leyenda = fila.get(ley_key, "-")
        if leyenda == "-":
            continue

        try:
            valor = float(fila.get(valor_key))
        except:
            valor = None

        if valor is not None:
            historial.append({"valor": valor, "tipo": leyenda, "fecha": fecha})

    # La medición más reciente se dibuja siempre como punto propio.
    if filas:
        try:
            valor_actual = float(filas[0].get(valor_key))
        except:
            valor_actual = None

        if valor_actual is not None:
            punto_actual = {
                "fecha_grafico": filas[0].get("fecha"),
                grafico_key: valor_actual
            }

    def agregar_grupos(segmento, tamano_grupo):
        i = 0
        while i < len(segmento):
            grupo = segmento[i:i + tamano_grupo]
            valores = []
            for fila in grupo:
                try:
                    valores.append(float(fila.get(valor_key)))
                except:
                    pass

            if valores:
                medio = len(grupo) // 2
                datos.append({
                    "fecha_grafico": grupo[medio].get("fecha"),
                    grafico_key: round(sum(valores) / len(valores), 4)
                })

            i += tamano_grupo

    # En alarma se agrupa de a 6 para suavizar la ráfaga de mediciones rápidas.
    # En estado normal se dibuja cada medición, que en el ESP llega cada 5 min.
    alarm_keys = ("ley_T", "ley_H", "ley_I", "ley_O")
    filas_cronologicas = list(reversed(filas))
    segmento = []
    estado_segmento = None
    for fila in filas_cronologicas:
        estado_fila = any(fila.get(key, "-") != "-" for key in alarm_keys)
        if estado_segmento is None:
            estado_segmento = estado_fila
        elif estado_fila != estado_segmento:
            agregar_grupos(segmento, 6 if estado_segmento else 1)
            segmento = []
            estado_segmento = estado_fila

        segmento.append(fila)

    if segmento:
        agregar_grupos(segmento, 6 if estado_segmento else 1)

    if punto_actual is not None and (
        not datos or datos[-1].get("fecha_grafico") != punto_actual["fecha_grafico"]
    ):
        datos.append(punto_actual)

    return datos, historial

@app.route("/datos")
def obtener_datos_csv():
    resultado_final = {}
    os.makedirs(DATA_DIR, exist_ok=True)
    csv_paths = sorted(glob.glob(os.path.join(DATA_DIR, "incubadora_*.csv")))

    for path in csv_paths:
        try:
            # extraer número real del archivo
            filename = os.path.basename(path)  # ej: incubadora_7.csv
            match = re.search(r"incubadora_(\d+)\.csv", filename)
            if not match:
                continue
            esp_num = int(match.group(1))
            esp_id = f"incubadora_{esp_num}"

            filas = read_csv_rows(path, MAX_ROWS)
            if not filas:
                continue

            # estructuras de salida
            datos_T, historial_T = construir_datos_parametro(
                filas, "temperatura", "ley_T", "temperatura_grafico"
            )
            datos_H, historial_H = construir_datos_parametro(
                filas, "humedad", "ley_H", "humedad_grafico"
            )
            datos_I, historial_I = construir_datos_parametro(
                filas, "iluminancia", "ley_I", "iluminancia_grafico"
            )
            datos_O, historial_O = construir_datos_parametro(
                filas, "oxigeno", "ley_O", "oxigeno_grafico"
            )

            # valores actuales
            top = filas[0]
            try: Tact = float(top.get("temperatura"))
            except: Tact = None
            try: Hact = float(top.get("humedad"))
            except: Hact = None
            try: Iact = float(top.get("iluminancia"))
            except: Iact = None
            try: Oact = float(top.get("oxigeno"))
            except: Oact = None
            ley_Tact = top.get("ley_T", "-")
            ley_Hact = top.get("ley_H", "-")
            ley_Iact = top.get("ley_I", "-")
            ley_Oact = top.get("ley_O", "-")

            estad = {
                "temperatura": calcular_estadisticas([r.get("temperatura") for r in filas]),
                "humedad": calcular_estadisticas([r.get("humedad") for r in filas]),
                "iluminancia": calcular_estadisticas([r.get("iluminancia") for r in filas]),
                "oxigeno": calcular_estadisticas([r.get("oxigeno") for r in filas])
            }

            ultimo_timestamp = None
            if filas:
                try:
                    ultimo_timestamp = filas[0].get("fecha")  # asumo formato datetime string
                except:
                    pass

            resultado_final[esp_id] = {
                "T_actual": Tact,
                "H_actual": Hact,
                "I_actual": Iact,
                "O_actual": Oact,
                "ley_T_actual": ley_Tact,
                "ley_H_actual": ley_Hact,
                "ley_I_actual": ley_Iact,
                "ley_O_actual": ley_Oact,
                "datos_T": datos_T,
                "datos_H": datos_H,
                "datos_I": datos_I,
                "datos_O": datos_O,
                "historial_T": historial_T,
                "historial_H": historial_H,
                "historial_I": historial_I,
                "historial_O": historial_O,
                "estadisticas": estad,
                "ultimo_timestamp": ultimo_timestamp   # 👈 agregado
            }

        except Exception as e:
            print(f"Error procesando {path}: {e}")

    # 👇 ordenar las claves numéricamente antes de devolver
    ordenado = dict(sorted(
        resultado_final.items(),
        key=lambda kv: int(kv[0].split("_")[1])
    ))

    return jsonify(ordenado)

@app.route("/download_xlsx")
def download_xlsx():
    """
    Genera un Excel en memoria donde cada CSV 'incubadora_N.csv' es una hoja.
    Intenta usar pandas; si no está instalado, hace fallback con openpyxl.
    """
    # buscar CSVs
    csv_dir = DATA_DIR
    csv_paths = sorted(glob.glob(os.path.join(csv_dir, "incubadora_*.csv")))
    if not csv_paths:
        return jsonify({"error": "no csv files found"}), 404

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"incubadoras_{ts}.xlsx"

    # intentar con pandas (más simple y preserva headers)
    try:
        import pandas as pd
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for path in csv_paths:
                sheet_name = os.path.splitext(os.path.basename(path))[0]
                # sheets names can't exceed 31 chars
                sheet_name = sheet_name[:31]
                try:
                    df = pd.read_csv(path, encoding='utf-8')
                except Exception:
                    # si falla leer con utf-8, intentar latin-1
                    df = pd.read_csv(path, encoding='latin-1')
                # escribir hoja
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        output.seek(0)
        return send_file(output,
                         as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e_pandas:
        # fallback: usar openpyxl directamente (sin pandas)
        try:
            from openpyxl import Workbook
            wb = Workbook()
            # eliminar sheet por defecto si lo tiene
            if wb.sheetnames:
                std = wb[wb.sheetnames[0]]
                wb.remove(std)
            for path in csv_paths:
                sheet_name = os.path.splitext(os.path.basename(path))[0][:31]
                ws = wb.create_sheet(title=sheet_name)
                # leer csv y volcar filas
                with open(path, newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for r_idx, row in enumerate(reader, start=1):
                        for c_idx, cell in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=cell)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(bio,
                             as_attachment=True,
                             download_name=filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e_openpyxl:
            # si todo falla, devolver error
            print("Error generando xlsx:", e_pandas, e_openpyxl)
            return jsonify({"error": "failed to generate xlsx"}), 500
        
if __name__ == "__main__":
    print("✅ Servidor Flask (CSV completo + ordenado) iniciado en http://localhost:5000/datos")
    app.run(host="0.0.0.0", port=5000, debug=True)
